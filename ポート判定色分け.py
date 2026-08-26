import re, unicodedata, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

SRC='元ファイル.xlsx'   # 入力：ポート確認シート（2月現調時／2016年ポート構成表／リンクアップダウン）
OUT='ポート確認_リンクアップダウン_色分け.xlsx'
JP='游ゴシック'

wb=openpyxl.load_workbook(SRC); ws=wb['Sheet1']
BLOCKS=[('コアスイッチ',1,51),('フロアスイッチ１F',5,27),('フロアスイッチ2F',9,27),
        ('フロアスイッチ3F',13,27),('フロアスイッチ4F',17,27)]

def norm(v):
    if v is None: return None
    s=unicodedata.normalize('NFKC',str(v)).strip()
    if s in ('','-','ー','―','‐','—'): return None
    s=s.replace(' ','').replace('　','').lower()
    m=re.fullmatch(r'(\d+)f?fe0?/(\d+)',s)
    if m: return str(int(m.group(2)))
    if re.fullmatch(r'\d+',s): return str(int(s))
    return s

def name_state(feb,old):
    a,b=norm(feb),norm(old)
    if a is None and b is None: return 'NONE'
    if a is None: return 'ONLY_OLD'
    if b is None: return 'ONLY_FEB'
    return 'MATCH' if a==b else 'DIFF'

def link_state(v):
    s='' if v is None else unicodedata.normalize('NFKC',str(v)).strip()
    if s in ('〇','○','o','O','0'): return 'UP'
    if s in ('×','x','X','✕'):      return 'DOWN'
    if s=='△':                      return 'HALF'
    return 'OTHER'

def judge(ns,ls):
    if ls=='HALF':
        return 'LINKDOWN','ケーブルは接続済みだがリンクダウン'
    if ls=='UP':
        return {'MATCH':('OK','問題なし（タグ名一致・リンクアップ）'),
                'NONE':('UNDOC','両表とも記載なしだがリンクアップ（台帳漏れ）'),
                'DIFF':('RENAME','タグ名不一致（名称変更の可能性）'),
                'ONLY_FEB':('RENAME','2016年表に記載なし（増設の可能性）'),
                'ONLY_OLD':('RENAME','2月現調時に記載なし')}[ns]
    if ls=='DOWN':
        return {'NONE':('UNUSED','未使用ポート（両表とも記載なし）'),
                'MATCH':('GONE','両表に記載ありだがリンクダウン（撤去済み？）'),
                'DIFF':('CHECK','要調査（タグ名不一致・リンクダウン）'),
                'ONLY_FEB':('CHECK','要調査（2月現調時のみ記載・リンクダウン）'),
                'ONLY_OLD':('CHECK','要調査（2016年表のみ記載・リンクダウン）')}[ns]
    return 'OTHER','判定不能（リンク欄が空欄など）'

STYLE={  # code: (fill, font色, 凡例ラベル)
 'OK':      ('C6EFCE','006100','問題なし：タグ名一致＋リンクアップ〇'),
 'RENAME':  ('FFEB9C','9C5700','タグ名要更新：タグ名不一致＋リンクアップ〇（名称変更／増設）'),
 'CHECK':   ('FFC7CE','9C0006','要調査：タグ名不一致＋リンクダウン×（「何？」の対象）'),
 'LINKDOWN':('F8CBAD','833C0C','リンクダウン△：ケーブルは刺さっているがリンクしていない'),
 'GONE':    ('E4D7F5','5B2C87','撤去済み？：タグ名一致だがリンクダウン×'),
 'UNDOC':   ('BDD7EE','1F4E79','台帳漏れ：両表とも記載なしだがリンクアップ〇'),
 'UNUSED':  ('D9D9D9','7F7F7F','未使用ポート：両表とも「-」＋リンクダウン×'),
 'OTHER':   ('FFFFFF','000000','判定不能'),
}
ORDER=['OK','RENAME','CHECK','LINKDOWN','GONE','UNDOC','UNUSED','OTHER']

thin=Side(style='thin',color='BFBFBF')
box=Border(left=thin,right=thin,top=thin,bottom=thin)

rows=[]
for bname,c0,last in BLOCKS:
    for r in range(4,last+1):
        port=ws.cell(r,c0).value
        if port is None: continue
        feb=ws.cell(r,c0+1).value; old=ws.cell(r,c0+2).value; lk=ws.cell(r,c0+3).value
        code,reason=judge(name_state(feb,old),link_state(lk))
        rows.append(dict(block=bname,row=r,c0=c0,port=port,feb=feb,old=old,link=lk,
                         code=code,reason=reason))
        bg,fg,_=STYLE[code]
        fill=PatternFill('solid',fgColor=bg)
        for c in range(c0,c0+4):
            cell=ws.cell(r,c)
            cell.fill=fill
            cell.font=Font(name=JP,sz=11,color=fg)
            cell.border=box
            cell.alignment=Alignment(horizontal='center' if c in (c0,c0+3) else 'left',
                                     vertical='center')

counts=Counter(r['code'] for r in rows)

# ---- ヘッダ装飾 ----
ws['A1'].font=Font(name=JP,sz=14,b=True)
hdr_fill=PatternFill('solid',fgColor='44546A')
for bname,c0,last in BLOCKS:
    t=ws.cell(2,c0); t.font=Font(name=JP,sz=12,b=True,color='1F4E79')
    for c in range(c0,c0+4):
        h=ws.cell(3,c)
        h.fill=hdr_fill; h.font=Font(name=JP,sz=11,b=True,color='FFFFFF')
        h.border=box; h.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
ws.row_dimensions[3].height=32
ws.freeze_panes='A4'

# ---- 凡例（本表の下） ----
L=53
ws.cell(L,1,'■ 凡例（色分けの意味）').font=Font(name=JP,sz=12,b=True)
for i,code in enumerate(ORDER):
    bg,fg,label=STYLE[code]
    r=L+1+i
    c=ws.cell(r,1); c.fill=PatternFill('solid',fgColor=bg); c.border=box
    c.value='　'
    d=ws.cell(r,2,label); d.font=Font(name=JP,sz=11,color=fg)
    e=ws.cell(r,3,f'{counts.get(code,0)}件')
    e.font=Font(name=JP,sz=11,b=True); e.alignment=Alignment(horizontal='left')
NOTE=L+1+len(ORDER)+1
notes=[
 '※ タグ名の比較ルール（前提）：全角/半角・空白を無視して比較。「1F FE0/01」形式と数字のみの記載は',
 '　 同じポート番号として扱い一致とみなす。「-」および空欄は「記載なし」として扱う。',
 '※ コアスイッチの「2016年ポート構成表」列でポート番号だけが記載されている行（1,2,3…）は、',
 '　 2月現調時の機器名と別物として「不一致」判定になる。実際は台帳に機器名が未記載なだけの可能性あり。',
 '※ △（リンクダウン）はタグ名の一致・不一致に関わらず橙色で表示（優先度が高いため）。',
 '※ 判定の内訳は「判定」シートを参照（オートフィルタで絞り込み可）。件数は作成時点のスナップショット。',
]
for i,t in enumerate(notes):
    ws.cell(NOTE+i,1,t).font=Font(name=JP,sz=10,color='595959')

# ---- 判定シート ----
if '判定' in wb.sheetnames: del wb['判定']
js=wb.create_sheet('判定')
js['A1']='判定サマリ'; js['A1'].font=Font(name=JP,sz=14,b=True)
js['A2']='区分'; js['B2']='件数'; js['C2']='内容'
for c in ('A2','B2','C2'):
    js[c].fill=hdr_fill; js[c].font=Font(name=JP,sz=11,b=True,color='FFFFFF'); js[c].border=box
for i,code in enumerate(ORDER):
    bg,fg,label=STYLE[code]; r=3+i
    a=js.cell(r,1,code); a.fill=PatternFill('solid',fgColor=bg)
    a.font=Font(name=JP,sz=11,b=True,color=fg); a.border=box
    b=js.cell(r,2,counts.get(code,0)); b.font=Font(name=JP,sz=11); b.border=box
    b.alignment=Alignment(horizontal='center')
    c=js.cell(r,3,label); c.font=Font(name=JP,sz=11,color=fg); c.border=box
js.cell(3+len(ORDER),1,'合計').font=Font(name=JP,sz=11,b=True)
tot=js.cell(3+len(ORDER),2,sum(counts.values())); tot.font=Font(name=JP,sz=11,b=True)
tot.alignment=Alignment(horizontal='center')
js.cell(3+len(ORDER)+1,1,'※ 件数は本ファイル作成時点のスナップショット。データを更新した場合はオートフィルタで再集計してください。').font=Font(name=JP,sz=10,color='595959')

HR=12
heads=['スイッチ','ポート','2月現調時','2016年ポート構成表','リンク','区分','判定理由','元セル']
for j,h in enumerate(heads,start=1):
    c=js.cell(HR,j,h); c.fill=hdr_fill
    c.font=Font(name=JP,sz=11,b=True,color='FFFFFF'); c.border=box
    c.alignment=Alignment(horizontal='center',vertical='center')
for i,d in enumerate(rows):
    r=HR+1+i; bg,fg,_=STYLE[d['code']]
    fill=PatternFill('solid',fgColor=bg)
    vals=[d['block'],d['port'],d['feb'],d['old'],d['link'],d['code'],d['reason'],
          f"Sheet1!{get_column_letter(d['c0'])}{d['row']}"]
    for j,v in enumerate(vals,start=1):
        c=js.cell(r,j,v); c.fill=fill; c.font=Font(name=JP,sz=11,color=fg); c.border=box
        c.alignment=Alignment(horizontal='center' if j in (2,5,6) else 'left',vertical='center')
js.auto_filter.ref=f'A{HR}:H{HR+len(rows)}'
js.freeze_panes=f'A{HR+1}'
for col,w in zip('ABCDEFGH',[16,10,30,30,8,11,44,14]):
    js.column_dimensions[col].width=w

wb.save(OUT)
print('saved',OUT)
print(Counter(r['code'] for r in rows))
